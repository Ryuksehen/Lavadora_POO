# Librería para trabajar con el sistema de archivos
# Permite crear carpetas y manejar rutas
import os
from datetime import datetime

# Librería para crear y editar archivos de Excel
from openpyxl import Workbook, load_workbook


# REPORTES
class Reportes:

    # Nombre de la carpeta donde se guardarán los reportes
    CARPETA = "reportes"

    # CREAR CARPETA DE REPORTES
    @staticmethod
    def crear_carpeta():

        if not os.path.exists(Reportes.CARPETA):

            os.makedirs(Reportes.CARPETA)


    # GENERAR FACTURA DEL CLIENTE (TXT)
    @staticmethod
    def generar_factura_txt(datos):

        # Verifica que exista la carpeta
        Reportes.crear_carpeta()

        # Construye la ruta del archivo
        archivo = os.path.join(Reportes.CARPETA, "factura.txt")

        # Abre el archivo en modo escritura
        with open(archivo, "w", encoding="utf-8") as f:

            # Encabezado de la factura
            f.write("========== LAVA SMART ==========\n")
            f.write("         COMPROBANTE\n\n")

            # Datos del cliente
            f.write(f"Cliente: {datos['cliente']}\n")
            f.write(f"Fecha: {datos['fecha']}\n\n")

            # Información del servicio
            f.write(f"Kilos lavados: {datos['kilos']}\n")
            f.write(f"Tipo de prenda: {datos['tipo_ropa']}\n")
            f.write(f"Método de lavado: {datos['metodo']}\n\n")

            # Costos
            f.write(f"Costo por kilo: ${datos['precio_kilo']}\n")
            f.write(f"Costo sin IVA: ${datos['costo_base']}\n")

            # Si hay prendas especiales se aplica recargo
            if datos["recargo"] > 0:

                f.write(f"Recargo prenda especial (5%): ${datos['recargo']}\n")

            # IVA del servicio
            f.write(f"IVA (19%): ${datos['iva']}\n\n")

            # Total final
            f.write(f"TOTAL A PAGAR: ${datos['total']}\n\n")

            # Mensaje final
            f.write("Gracias por usar Lava Smart\n")

        # Mensaje en consola
        print("✔ factura.txt generada")


    # GENERAR REPORTE PARA ADMINISTRADOR
    @staticmethod
    def generar_excel(datos):

        # Asegura que exista la carpeta
        Reportes.crear_carpeta()

        # Ruta del archivo Excel
        archivo = os.path.join(Reportes.CARPETA, "reporte_admin.xlsx")


        # SI EL ARCHIVO YA EXISTE
        # Se abre para agregar más registros
        if os.path.exists(archivo):

            wb = load_workbook(archivo)
            ws = wb.active


        # SI EL ARCHIVO NO EXISTE
        # Se crea un nuevo archivo Excel
        else:

            wb = Workbook()
            ws = wb.active

            # Se agregan los encabezados de la tabla
            ws.append([
                "Cliente",
                "Fecha",
                "Kilos",
                "Tipo ropa",
                "Metodo",
                "Costo base",
                "IVA",
                "Total",
                "Ganancia",
                "Costo energia"
            ])

        # AGREGAR FILA DE DATOS
        ws.append([
            datos["cliente"],
            datos["fecha"],
            datos["kilos"],
            datos["tipo_ropa"],
            datos["metodo"],
            datos["costo_base"],
            datos["iva"],
            datos["total"],
            datos["ganancia"],
            datos["energia"]
        ])

        # Guarda los cambios en el archivo
        wb.save(archivo)

        # Mensaje en consola
        print("✔ reporte_admin.xlsx actualizado")